# -*- coding:utf-8 -*-
import re
import openpyxl

def filtter():
    level = {
        0: "正常",
        1: "低",
        2: "中",
        3: "高"
    }
    big_categorize = {
        "刷单": {
            "A": ["平台", "推广", "小助手", "客服", "斗因", "抖音", "快手"],
            "B": ["点赞", "关注", "网红", "双击", "兼职"],
            "C": ["一单一", "结算", "押金", "收入", "每单", "不收取", "元", "佣金"]
        },
        "冒充售后": {
            "A": ["抖音", "中心", "商城"],
            "B": ["购买", "代理商", "通知到", "开通"],
            "C": ["扣除", "费用", "关闭"]
        },
        "冒充公检法_high": {
            "A": ["公安局", "户政科", "社保局", "管理局", "刑侦支队", "互联网举报中心", "药监局", "反欺诈中心", "拘留审查部", "110报案中心", "稽查科", "侦查队", "网信办", "监督局", "通信局", "税务稽查", "反电信", "疾控","互联网信息办公室"],
            "B": ["身份证", "协商文件", "护照", "大量发送", "举报", "协调"],
            "C": ["来一趟", "异常使用", "违规使用", "强制停机", "配合调查", "中奖短信", "核对", "核实", "协查", "批评政府", "通知"]
        },
        "冒充机关单位购物_high": {
            "A": ["疾控", "消防", "中队", "部队", "公安", "武警"],
            "B": ["猪肉", "猪热", "五花肉", "后腿肉", "药物"],
            "C": ["现金", "结算"]
        },
        "冒充机关退费_high": {
            "A": ["医保中心"],
            "B": ["灵活就业", "邻国就业"],
            "C": ["费"]
        },
        "快递退费_high": {
            "A": ["中通", "快递", "韵达", "天天", "圆通", "快手", "商城"],
            "B": ["弄丢", "遗失", "丢失", "丢了", "运单尾号", "不小心"],
            "C": ["商品价格", "多少钱", "协商", "理赔", "什么产品", "赔款", "退款", "退保", "钱退还","赔付"]
        },
        #金条等误开会员类
        "京东金条_high": {
            "A": ["京东", "金条", "白条", "淘宝", "快手", "抖音", "88V", "八八会员业务", "蘑菇街", "商城财务中心","微信商城", "百万保障"],
            "B": ["实名制", "注册", "身份证", "无意中", "不小心", "实习生", "实习员工", "操作失误", "开通", "总结算日","绑定"],
            "C": ["真心", "征信", "注销", "关闭", "交易记录", "扣费", "取消", "协商", "代理商", "抱歉", "扣款"]
        },
        "平台送礼_high": {
            "A": ["天猫超市", "美团", "商家联盟", "淘宝", "抖音", "快手", "新相应", "心相印", "新乡一"],
            "B": ["周年庆", "回馈", "派发", "助力"],
            "C": ["赠送", "免费", "礼品", "送礼"]
        },
    }

    def check_filter(keywords, text):
        rlist = re.findall("|".join(keywords), text)
        result = list(dict.fromkeys(rlist))
        return result, len(result)

    def get_level(key_info, text):
        last_list = []
        cout_list = []
        for k, v in key_info.items():
            tmp_list, tmp_count = check_filter(v, text)
            last_list.extend(tmp_list)
            cout_list.append(tmp_count)
        level_num = len([i for i in cout_list if i > 0])
        level_info = level[level_num]
        return level_info, level_num, last_list, cout_list

    workbook = openpyxl.load_workbook('content.xlsx')
    sheet = workbook.active
    max_column = sheet.max_column
    sheet.cell(row=1, column=max_column + 1, value="命中词")
    sheet.cell(row=1, column=max_column + 2, value="类别")
    sheet.cell(row=1, column=max_column + 3, value="等级")
    i = 0
    for row in sheet.iter_rows():
        for cell in row:
            if i > 0 and cell.column == 1:
                # 获取content内容
                text = cell.value
                final_last_list = []
                if not isinstance(text, str):
                    continue
                final_categorize_dict = {}
                level_num_dict = {}
                for categorize, key_info in big_categorize.items():
                    level_info, level_num, last_list, cout_list = get_level(key_info, text)
                    if level_num > 0:
                        final_last_list.extend(last_list)
                        #final_level_info += level_info + " "
                        #final_categorize += categorize + " "
                        sum_count = sum(cout_list)
                        if sum_count in final_categorize_dict:
                            final_categorize_dict[sum_count] = final_categorize_dict[sum_count] + " " + categorize
                        else:
                            final_categorize_dict[sum_count] = categorize
                        level_num_dict[level_info] = level_num
                        max_key = max(final_categorize_dict)
                        final_categorize = final_categorize_dict.get(max_key)
                        final_level_info = max(level_num_dict, key=level_num_dict.get)
                    else:
                        final_categorize = ""
                        final_level_info = level_info
                sheet.cell(row=i + 1, column=max_column + 1, value=",".join(list(set(final_last_list))))
                sheet.cell(row=i + 1, column=max_column + 2, value=final_categorize)
                sheet.cell(row=i + 1, column=max_column + 3, value=final_level_info)
        i += 1
    workbook.save("hit_result.xlsx")

filtter()